from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpRequest
from django.shortcuts import render, redirect, get_object_or_404


from  app.form_pm import *
from app.decorators import allowed_users,manger_and_viewer_only,manger_only,staff_admin_only,manger_and_viewer_engineer_only

from .filters import *

from django.core.exceptions import *

from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.db.models import Q, Sum

from urllib.parse import quote
import  app.user_access

from django import template
from django.contrib.auth.models import Group

from django.http import HttpResponse
from urllib import parse
from urllib.parse import urlencode
from django.core.paginator import Paginator

import pandas as pd
import os
from app.pm_doc_manager.pm_doc_builder import process_pm_doc_task,delete_pm_doc_task
import app.pm_doc_manager.pdf_file_directory_manager as fd_manager
import app.utility as util
import pandas as pd
import openpyxl
import app.pm_doc_manager.pm_export as exporter

from openpyxl.utils.dataframe import dataframe_to_rows
import calendar
from django.db.models import  Count

def run_process_pm_doc_task(request):
    try:

       process_pm_doc_task()

            #delete_pm_doc_task()

    except Exception as ex:
            print(str(ex))

    return redirect("manage_project")


def create_qstring(request):
    x_url = request.build_absolute_uri()
    qsl_dic = dict(parse.parse_qsl(parse.urlsplit(x_url).query))
    qsl_dic = {key: val for key, val in qsl_dic.items()}
    qstr = urlencode(qsl_dic)
    return qstr

def pm_checkEmpyQueryString(request):
    listQueryParma = request.GET

    listValInParma = dict(filter(lambda elem: len(elem[1]) > 0, listQueryParma.items()))
    isNotEmplyQuery = bool(listQueryParma) and len(listValInParma) > 0
    return isNotEmplyQuery



@login_required(login_url='login')
@manger_only
def summarize_project_pm(request):
    myuser = request.user
    today=date.date.today()

    isNotEmplyQuery = pm_checkEmpyQueryString(request)
    if isNotEmplyQuery:
      include_expired = request.GET.get('expired_project')
      if include_expired is None:
        my_all_pm = PreventiveMaintenance.objects.filter(project__company__manager__user=myuser,project__project_end__gte=today)
      else:
        my_all_pm = PreventiveMaintenance.objects.filter(project__company__manager__user=myuser)
    else:
        my_all_pm = PreventiveMaintenance.objects.filter(id=0)

    pmFilter = PMSummaryFilter(request.GET, request=request, queryset=my_all_pm)

    if pmFilter.qs.count() > 0:
        my_all_pm = pmFilter.qs
        listIDs = [x.id for x in my_all_pm]
        df = exporter.export_pm_summary_by_company_project(listIDs)
        if df.empty==False:
            if include_expired is None:
                return refactor2_report_summary(df, "search")
            else:
                return refactor2_report_summary(df, "search",True)
        else:
          context = {}
          return render(request, 'app/pm_export.html', context)

    else:
        my_all_pm = None
        context = {'PMSummaryFilter': pmFilter}
        return render(request, 'app/pm_summary.html', context)

@login_required(login_url='login')
@manger_only
def summarize_all(request):
    myuser = request.user
    today=date.date.today()
    my_all_pm = PreventiveMaintenance.objects.filter(project__company__manager__user=myuser,
                                                     project__project_end__gte=today)
    listIDs = [x.id for x in my_all_pm]
    df = exporter.export_pm_summary_by_company_project(listIDs)
    if df.empty==False:
       return refactor2_report_summary(df,"all")
    else:
        context = {}
        return render(request, 'app/pm_export.html', context)


def refactor2_report_summary(df,doc_type,includeExpiration=False):
    if doc_type == 'search':
        if includeExpiration:
         file_name = f"Project_PM_ByCompany_IncludeExpiration.xlsx"
        else:
            file_name = f"Project_PM_ByCompany.xlsx"
    else:
        file_name = f"All_Project_PM.xlsx"

    df = df.fillna(value='')
    df = df.reset_index(drop=False)
    df.insert(0, "No", df["index"] + 1, True)
    df = df.drop(columns=["index"])

    col_remove = 'ID'
    if col_remove in df.columns:
        df = df.drop([col_remove], axis=1)

    if len(df) > 0:
        try:
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet()
            for r in dataframe_to_rows(df, index=False, header=True, ):
                ws.append(r)
        except openpyxl.utils.exceptions.IllegalCharacterError as e:
            raise e

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={file_name}'
    wb.save(response)

    return response

@login_required(login_url='login')
@manger_and_viewer_engineer_only
def report_pm(request):
    myuser = request.user
    isNotEmplyQuery = pm_checkEmpyQueryString(request)

    xtoday=date.date.today()
    first_d,last_d  = calendar.monthrange(xtoday.year, xtoday.month)
    first_of_month=datetime.datetime(year=xtoday.year,month=xtoday.month,day=1)
    last_of_month=datetime.datetime(year=xtoday.year,month=xtoday.month,day=last_d)

    str_FirstMonth=  first_of_month.strftime("%Y-%m-%d")
    str_LastMonth = last_of_month.strftime("%Y-%m-%d")

    # if isNotEmplyQuery:
    #     my_all_pm = PreventiveMaintenance.objects.filter(project__company__manager__user=myuser).order_by(
    #         '-planned_date')
    # else:
    #     my_all_pm = PreventiveMaintenance.objects.filter(project__company__manager__user=myuser,
    #                                                      planned_date__gte=str_FirstMonth,
    #                                                      planned_date__lte=str_LastMonth).order_by('-planned_date')

    if isNotEmplyQuery:
        my_all_pm = PreventiveMaintenance.objects.filter(project__company__manager__user=myuser).order_by('-planned_date')
        if my_all_pm.count() ==0: # if site mange ==0 , we chahnge to search on engineer
         my_all_pm = PreventiveMaintenance.objects.filter(project__company__engineer__user=myuser) .order_by('-planned_date')
    else:
        my_all_pm = PreventiveMaintenance.objects.filter(project__company__manager__user=myuser ,
                                                         planned_date__gte=str_FirstMonth,planned_date__lte=str_LastMonth).order_by('-planned_date')
        if my_all_pm.count()==0:
            my_all_pm = PreventiveMaintenance.objects.filter(project__company__engineer__user=myuser,
                                                             planned_date__gte=str_FirstMonth,
                                                             planned_date__lte=str_LastMonth).order_by('-planned_date')

    pmFilter = PMFilter(request.GET, request=request, queryset=my_all_pm)
    if pmFilter .qs.count() > 0:
        my_all_pm = pmFilter.qs
        listIDs = [x.id for x in my_all_pm]
        request.session['query_pm_plan'] = listIDs
    else:
        my_all_pm = None


    date_from = request.GET.get('planned_date__gt')
    date_to = request.GET.get('planned_date__lt')
    date_between_qstring = ''
    if (date_from is not None) and (date_to is not None):
        date_between_qstring = f"?date_from={date_from}&date_to={date_to}"

    context = {'PMFilter': pmFilter,'pmList':my_all_pm,'isNotEmplyQuery':isNotEmplyQuery,
               'date_between_qstring': date_between_qstring}

    return render(request, 'app/pm_report.html', context)


@login_required(login_url='login')
@manger_and_viewer_engineer_only
def export_pm_plan(request):
    response = export_pm_refactor(request, 'query_pm_plan','pm')
    return response
@login_required(login_url='login')
@manger_and_viewer_engineer_only
def export_pm_item(request):
    response = export_pm_refactor(request, 'query_pm_plan','item')
    return response
def export_pm_refactor(request, query_session,doc_type):

    listIDs = request.session.get(query_session, None)

    if listIDs != None:
        date_btw = ''
        buildtime = datetime.datetime.now().strftime('%d%m%y_%H%M')

        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        if date_from is not None and date_to is not None:
            date_btw = f'_From{date_from}_To{date_to}'

        if doc_type == 'pm':
            df = exporter.export_pm_plan(listIDs)
            if df.empty:
                messages.error(request, "No PM Plane during the given period")
                return render(request, 'app/pm_export.html', {})
            file_name = f"PM-Plan_{date_btw}_{buildtime}.xlsx"
        else: # doc_type=='item'
            # get item_id from listIDs
            itemIDs=[ obj.id  for obj in PM_Inventory.objects.filter(pm_master__in=listIDs) ]
            df = exporter.export_pm_iventory_item(itemIDs)
            if df.empty:
                messages.error(request, "No PM Items return")
                return render(request, 'app/pm_export.html', {})

            file_name = f"PM-Item_{date_btw}_{buildtime}.xlsx"


        df = df.fillna(value='')
        df = df.reset_index(drop=False)
        df.insert(0, "No", df["index"]+1,True)
        df=df.drop(columns=["index"])


        col_remove = 'ID'
        if col_remove in df.columns:
            df = df.drop([col_remove], axis=1)

        if len(df) > 0:
            try:
                wb = openpyxl.Workbook(write_only=True)
                ws = wb.create_sheet()
                for r in dataframe_to_rows(df, index=False, header=True, ):
                    ws.append(r)
            except openpyxl.utils.exceptions.IllegalCharacterError as e:
                raise e

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'

        wb.save(response)

        return response

        context = {}
        return render(request, 'app/pm_export.html', context)



@login_required(login_url='login')
@manger_and_viewer_engineer_only
def build_weasyprint_pm_doc_pdf(request,id=0):

    task_obj=TaskSchedule_PMDoc.objects.create(created_date=datetime.datetime.now(),status=0,owner=request.user,pm_id=id)

    return render(request, 'app/pm_complete_report.html')

    # overview_file_path=f"pm_overview_report.pdf"
    #pdf_bytes = xyz_pdf.create_pdf(report_data, overview_dict, None)
    # response = HttpResponse(pdf_bytes,content_type='application/pdf')
    # response['Content-Disposition'] = f'attachment; filename="{overview_file_path}"'
    # return response

@login_required(login_url='login')
@manger_and_viewer_engineer_only
def filter_pmItem(request, pmItemList):

    listBrandIDByPMItems=list( pmItemList.order_by('inventory__brand_id').values_list('inventory__brand_id',flat=True).distinct())
    brandByPMItems=Brand.objects.filter(id__in=listBrandIDByPMItems)
    request.session['listBrandIDByPMItems']=listBrandIDByPMItems

    pmItemFilter = PMInventoryItemFilter(request.GET, request=request, queryset=pmItemList)

    pmItemList=pmItemFilter.qs


    del request.session['listBrandIDByPMItems']

    return pmItemFilter, pmItemList
@login_required(login_url='login')
@manger_and_viewer_only
def copy_pm_inventory(request,pm_id):

    msg_check=app.user_access.check_user_to_do(request,pm_id,"CopyPM")
    if msg_check is not None: return HttpResponse(msg_check)

    # pm is used to copy the inventory
    pm_obj = get_object_or_404(PreventiveMaintenance, pk=pm_id)

    # new pm
    form = PM_MasterForm()
    listItem = pm_obj.pm_inventory_set.filter()

    if request.method == "POST":
        form = PM_MasterForm(request.POST)
        new_pm_obj = form.save(commit=False)

        new_pm_obj.project_id = pm_obj.project_id
        new_pm_obj.save()

        # way insert bulk , copy all item from master including  no-pm
        listNewCopiedItems=[ PM_Inventory(pm_master=new_pm_obj,is_pm=item.is_pm,inventory=item.inventory,) for item in listItem ]

        if len(listNewCopiedItems)>0:
                  PM_Inventory.objects.bulk_create( listNewCopiedItems) 
        # way insert each          
        # for item in listItem:
        #     item.id=None
        #     item.pm_master_id=new_pm_obj.id
        #     item.save()
        return redirect('manage_pm', project_id=pm_obj.project.id, id=0)
    else:
        listSubComp=SubCompany.objects.filter( head_company_id=pm_obj.project.company.id  )
        form.fields['customer_company'].queryset=listSubComp
        form.fields['customer_company'].value=pm_obj.customer_company.id
        form.fields['team_lead'].value=pm_obj.team_lead
        form.fields['engineer'].value = pm_obj.engineer
        form.fields['contact_name'].value = pm_obj.contact_name
        form.fields['contact_telephone'].value = pm_obj.contact_telephone


    context = {'pmInfo':pm_obj,'form': form,'inventoryPMList':listItem}
    return render(request, "app/pm_copy_inventory.html", context)

@login_required(login_url='login')
@manger_and_viewer_engineer_only
def update_pm_inventory(request,pm_id,id=0):

    msg_check=app.user_access.check_user_to_do(request,pm_id,"UpdatePMItem")
    if msg_check is not None: return HttpResponse(msg_check)

    pm_obj = get_object_or_404(PreventiveMaintenance, pk=pm_id)
    pmItemFilter, pmItemList=filter_pmItem(request,pm_obj.pm_inventory_set)

    searched_item_str='All'
    if len(request.GET)>0:
        brand_id=int(request.GET.get("brand"))
        brand_obj = get_object_or_404(Brand, pk=brand_id)
        brand_name=brand_obj.brand_name
        is_pm = request.GET.get("is_pm")
        searched_item_str=f"Brand={brand_name}"
        if is_pm!='unknown':
            searched_item_str = f" {searched_item_str} | Is PM={is_pm}"

    noTotalPMInventory = pmItemList.filter(is_pm=True).count()
    # defined template
    # inventoryPMList=pm_obj.pm_inventory_set.filter(inventory__pm_inventory_template__isnull=False)
    inventoryPMList = pmItemList.filter(inventory__pm_inventory_template__isnull=False,is_pm=True)
    # undefine template
    # inventory_NoTemplatePMList = pm_obj.pm_inventory_set.filter(inventory__pm_inventory_template__isnull=True)
    inventory_NoTemplatePMList = pmItemList.filter(inventory__pm_inventory_template__isnull=True,is_pm=True)
    noTemplateFound=len(inventory_NoTemplatePMList)

    inventoryNoPMItemList =pmItemList.filter(is_pm=False)
    noInventoryNoPMItemList=len(inventoryNoPMItemList )


    item_obj = None
    if request.method == "GET":

        if id == 0:  # new detail
            if noTotalPMInventory >0:
             first_pm_item=pm_obj.pm_inventory_set.first()
             form =PM_InventoryForm(initial={'pm_master': pm_obj,'inventory':first_pm_item})
            else:
             form = PM_InventoryForm()

            form.fields['pm_engineer'].queryset = Employee.objects.filter(is_inactive=False)
            form.fields['document_engineer'].queryset = Employee.objects.filter(is_inactive=False)

        else:
            item_obj = get_object_or_404(PM_Inventory, pk=id)
            form = PM_InventoryForm(instance=item_obj)

            form.fields['pm_engineer'].queryset = Employee.objects.filter(Q(is_inactive=False) | Q(id=item_obj.pm_engineer_id))
            form.fields['document_engineer'].queryset = Employee.objects.filter(Q(is_inactive=False) | Q(id=item_obj.document_engineer_id))

    else:# post
        if id == 0:
            if len(pmItemList) > 0:  # is_pm=True
                if len(request.GET)>0 : # update only searched items
                    list_item=pmItemList.filter(is_pm=True)
                else: # update all items
                     list_item=PM_Inventory.objects.filter(pm_master_id=pm_id,is_pm=True)
                upatedItems=[]
                form = PM_InventoryForm(request.POST)
                if form.is_valid():
                    temp_obj = form.save(commit=False)
                    for item in list_item:
                        item.actual_date=temp_obj.actual_date
                        item.document_date=temp_obj.document_date

                        item.pm_engineer=temp_obj.pm_engineer
                        item.document_engineer=temp_obj.document_engineer

                        item.call_number=temp_obj.call_number
                        item.pm_document_number=temp_obj.pm_document_number

                        item.remark=temp_obj.remark
                        item.is_pm=temp_obj.is_pm

                        upatedItems.append(item)

                    colListToUpdate=  ['actual_date','document_date','pm_engineer','document_engineer'
                                   ,'call_number','pm_document_number','remark','is_pm']
                    PM_Inventory.objects.bulk_update(upatedItems,colListToUpdate)
            else:
                messages.info(request, f'No searched item to update for PM Inventory Item .')

        else:  # save from  edit
            item_obj = get_object_or_404(PM_Inventory, pk=id)
            form = PM_InventoryForm(request.POST, instance=item_obj)
            if form.is_valid():
                if form.has_changed():
                    form.save()
                    messages.success(request, f'PM Inventory Item has been updated successfully.')
                else:
                    messages.info(request, f'No any update for PM Inventory Item .')
            else:
                messages.error(request, form.errors)
        return redirect('update_pm_inventory', pm_id=pm_id, id=0)

    if len(pmItemList)>0:
        has_item_to_update="False"
    else:
        has_item_to_update="True"
    context = {'pmInfo':pm_obj,'pmItemFilter':pmItemFilter,'searched_item_str':searched_item_str,
               'has_item_to_update':has_item_to_update,
               'inventoryPMList':inventoryPMList,
               'inventory_NoTemplatePMList': inventory_NoTemplatePMList,
               'noTemplateFound': noTemplateFound,
               'noTotalPMInventory':noTotalPMInventory,
               'inventoryNoPMItemList':inventoryNoPMItemList,
               'noinventoryNoPMItemList':noInventoryNoPMItemList,
               'form':form,
               'pm_item':item_obj,
               }
    return render(request, "app/pm_inventory.html", context)

@login_required(login_url='login')
@manger_and_viewer_only
def manage_pm(request, project_id, id=0):

    msg_check=app.user_access.check_user_to_do(request,project_id,"AddPM")
    if msg_check is not None: return HttpResponse(msg_check)

    project_obj = get_object_or_404(Project, pk=project_id)
    myuser = request.user

    if request.method == "GET":
        listSubComp=SubCompany.objects.filter( head_company_id=project_obj .company.id )
        if id == 0:  # new detail
            form = PM_MasterForm()
            form_mode = 'New'

            form.fields['team_lead'].queryset = Employee.objects.filter(is_inactive=False).order_by('-is_team_lead',                                                                                       'employee_name')
            form.fields['engineer'].queryset = Employee.objects.filter(is_inactive=False).order_by('employee_name')
        else:
            obj = get_object_or_404(PreventiveMaintenance, pk=id)
            form = PM_MasterForm(instance=obj)
            form_mode = 'Update'

            form.fields['team_lead'].queryset = Employee.objects.filter(Q(is_inactive=False) | Q(id=obj.team_lead_id)).order_by('-is_team_lead','employee_name')
            form.fields['engineer'].queryset = Employee.objects.filter(Q(is_inactive=False) | Q(id=obj.engineer_id)).order_by('employee_name')

        form.fields['customer_company'].queryset=listSubComp


    else:
        if id == 0:  # save from  new
            form = PM_MasterForm(request.POST)
            form_mode = 'New'
            list_inventory = Inventory.objects.filter(project_id=project_obj.id)
            if form.is_valid() and len(list_inventory)>0 :

                pm_master_obj = form.save(commit=False)
                pm_master_obj.project=project_obj
                pm_master_obj.save()

                list_pm_items=[ PM_Inventory(inventory=item,pm_master=pm_master_obj,is_pm=True) for item in list_inventory ]
                if len(list_pm_items)>0:
                  PM_Inventory.objects.bulk_create( list_pm_items) # batch_size=999
                messages.success(request, f'Create PM Plan and {len(list_inventory)} inventories successfully.')
            else:
                messages.error(request, form.errors)
        else:  # save from  edit
            obj = get_object_or_404(PreventiveMaintenance, pk=id)
            form =PM_MasterForm(request.POST, instance=obj)
            form_mode = 'Update'
            if form.is_valid():
                if form.has_changed():
                    form.save()
                    messages.success(request, f'PM Plan has been updated successfully.')
                else:
                    messages.info(request, f'No any update for PM Plan.')
            else:
                messages.error(request, form.errors)

        return redirect('manage_pm', project_id=project_id, id=0)

    pmList=PreventiveMaintenance.objects.filter(project_id=project_id)


    context = {'projectInfo': project_obj,'form': form,'mode':form_mode,
               'pmList':pmList
               }
    return render(request, "app/pm_master.html", context)

@login_required(login_url='login')
@manger_and_viewer_only
def delete_pm(request, id):

    msg_check=app.user_access.check_user_to_do(request,id,"DeletePM")
    if msg_check is not None: return HttpResponse(msg_check)

    x_obj = PreventiveMaintenance.objects.get(pk=id)
    if request.method == "GET":
        items=x_obj.pm_inventory_set.filter(is_pm=True)
        total_items = items.count()
        completed_items=items.filter(actual_date__isnull=False,document_date__isnull=False).count()
    try:
        if request.method == "POST":
            x_obj.delete()

    except Exception as ex:
        messages.ERROR(ex)
    if request.method == "GET":
        context = {'pm': x_obj, 'totalPMItems': total_items,
                   'completedPMItems': completed_items}
        return render(request, 'app/pm_delete.html', context)
    else:
        return redirect('manage_pm', project_id=x_obj.project.id, id=0)
    




