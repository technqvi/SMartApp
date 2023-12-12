from datetime import datetime, date
import calendar
from dateutil.relativedelta import relativedelta
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpRequest
from django.shortcuts import render, redirect, get_object_or_404

from .decorators import allowed_users, manger_and_viewer_only, manger_only, staff_admin_only
from .filters import *
from .forms import *
from django.core.exceptions import *

from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.db.models import Q, Sum

from urllib.parse import quote

import app.user_access
from django import template
from django.contrib.auth.models import Group

import xlsxwriter
import io
from django.http import HttpResponse
from urllib import parse
from urllib.parse import urlencode
from django.core.paginator import Paginator

import pandas as pd
import openpyxl

from django.core import serializers

from dataclasses import dataclass
from dataclasses_json import dataclass_json
import json
import operator
import re
import app.report_manager as reporter
import app.dashboard_manager as dashboard

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

@dataclass_json
@dataclass
class ImportedModel:
    brand_id: int
    brand_name: str
    model_name: str
    model_des: str
    note: str
    isValid: bool


def list_group(request):
    x_group = []
    for g in request.user.groups.all():
        x_group.append(g.name)
    return x_group


def check_group(request, group_name):
    return request.user.groups.filter(name=group_name).exists()


def is_valid_queryparam(param):
    return param != '' and param is not None


def checkEmpyQueryString(request):
    listQueryParma = request.GET

    listValInParma = dict(filter(lambda elem: len(elem[1]) > 0, listQueryParma.items()))
    isNotEmplyQuery = bool(listQueryParma) and len(listValInParma) > 0
    return isNotEmplyQuery


def create_qstring_without_page(request):
    x_url = request.build_absolute_uri()
    qsl_dic = dict(parse.parse_qsl(parse.urlsplit(x_url).query))

    qsl_dic = {key: val for key, val in qsl_dic.items() if key != settings.MY_PAGE_NAME}
    qstr = urlencode(qsl_dic)
    return qstr


def create_all_qstring(request):
    x_url = request.build_absolute_uri()
    qsl_dic = dict(parse.parse_qsl(parse.urlsplit(x_url).query))
    qstr = urlencode(qsl_dic)
    return qstr


def home(request):
    """Renders the home page."""
    assert isinstance(request, HttpRequest)

    countStatusList = None

    my_sites = []
    sites_dropdownlist = []

    current_date = datetime.date.today()
    no_year_lookup = settings.MY_YEAR_LOOKUP
    this_year = current_date.year
    my_dropdownlist = []
    for i in range(0, no_year_lookup + 1):
        my_dropdownlist.append(this_year - i)
    title_time = f"{this_year}"

    title_site = "All Sites"
    site_in_control = False

    incident_date_from = None
    incident_date_end = None
    siteID = None

    if request.user.is_anonymous == False:

        my_sites_qs = Company.objects.filter(manager__user=request.user)
        my_sites = list(my_sites_qs.values("id", "company_name", "company_full_name"))
        listSiteID = list([site['id'] for site in my_sites])

        if len(listSiteID) > 0:

            site_in_control = True
            # create dropdowlist
            if len(my_sites) > 0:
                sites_dropdownlist = my_sites.copy()
                sites_dropdownlist.insert(0, {"id": 0, "company_name": "All", "company_full_name": 'Sites'})

            # first loading for site
            select_company = request.GET.get('select_company')
            if select_company is not None:
                if int(select_company) > 0:
                    listSiteID = [select_company]
                    title_site = [f"{site['company_name']}-{site['company_full_name']}" for site in my_sites if
                                  site['id'] == int(select_company)][0]

            # first loading for time
            select_month = request.GET.get('select_month')
            select_year = request.GET.get('select_year')
            if select_month is not None:
                if int(select_month) == 0:
                    incident_date_from = datetime.datetime(int(select_year), 1, 1)
                    incident_date_end = incident_date_from + relativedelta(years=1)
                    title_time = f"{select_year}"
                else:
                    incident_date_from = datetime.datetime(int(select_year), int(select_month), 1)
                    incident_date_end = incident_date_from + relativedelta(months=1)
                    title_time = f'{incident_date_from.strftime("%B")}-{incident_date_from.strftime("%Y")}'

            else:
                if select_year is not None:
                    incident_date_from = datetime.datetime(int(select_year), 1, 1)
                    title_time = f"{select_year}"
                else:
                    incident_date_from = datetime.datetime(int(this_year), 1, 1)

                incident_date_end = incident_date_from + relativedelta(years=1)

            # list incident
            count_status_df = dashboard.count_incident(listSiteID, incident_date_from, incident_date_end)
            # if select site then convert datatime to date and get site for link to incident

            # if len(listSiteID) == 1 and int(select_month) > 0:
            incident_date_from = incident_date_from.date().strftime('%Y-%m-%d')
            incident_date_end = incident_date_end + relativedelta(days=-1)
            incident_date_end = incident_date_end.date().strftime('%Y-%m-%d')

            if len(listSiteID) == 1:
                siteID = listSiteID[0]

            if count_status_df.empty == False:

                # listAllStatus = [s['incident_status_name'] for s in
                #                  Incident_Status.objects.filter(~Q(id=settings.INCIDENT_CODE_CANCELED)).values("incident_status_name")]
                listAllStatus = [{'id': s['id'], 'incident_status_name': s['incident_status_name']} for s in
                                 Incident_Status.objects.filter(~Q(id=settings.INCIDENT_CODE_CANCELED)).values(
                                     "id", "incident_status_name")]

                for status in listAllStatus:
                    status_id = status['id']
                    df_tem = count_status_df.query('id==@status_id')
                    if df_tem.empty:
                        count_status_df = count_status_df.append(
                            {'id': status['id'], 'incident_status_name': status['incident_status_name'], 'count': 0},
                            ignore_index=True)

                sum_status_df = pd.DataFrame(data={'id': [0], 'incident_status_name': ['All Tickets'],
                                                   'count': [count_status_df['count'].sum()]})
                count_status_df = pd.concat([sum_status_df, count_status_df])
                count_status_df.reset_index(drop=True, inplace=True)
                countStatusList = count_status_df.to_dict('records')
            else:  # All Asset =0 due to emplty dataframe
                countStatusList = [{'id': '', 'incident_status_name': 'All Tickets', 'count': 0}]

    return render(request,
                  'app/index.html',
                  {
                      'site_in_control': site_in_control,
                      'countStatusList': countStatusList,
                      'current_year': datetime.datetime.today().year,

                      'my_sites': sites_dropdownlist,
                      'my_years': my_dropdownlist,
                      'title_site': title_site,
                      'title_time': title_time,
                      'incident_date_from': incident_date_from,
                      'incident_date_end': incident_date_end,
                      'siteID': siteID

                  })


def filter_project(request):
    myuser = request.user
    isNotEmplyQuery = checkEmpyQueryString(request)

    if isNotEmplyQuery:
        include_expired = request.GET.get('expired_project')
        if include_expired is None:
            my_all = Project.objects.filter(company__manager__user=myuser, project_end__gte=datetime.date.today(),
                                            is_dummy=False).order_by('-id')
        else:
            my_all = Project.objects.filter(company__manager__user=myuser,
                                            is_dummy=False).order_by('-id')
    else:
        my_all = Project.objects.filter(company__manager__user=myuser, project_end__gte=datetime.date.today(),
                                        is_dummy=False).order_by('-id')[
                 :settings.MY_PAGE_PER]

    xFilter = ProjectFilter(request.GET, request=request, queryset=my_all)
    if xFilter.qs.count() > 0:
        projectList = xFilter.qs
        listIDs = [x.id for x in projectList]
        request.session['query_project'] = listIDs

        paginator = Paginator(projectList, settings.MY_PAGE_PER)
        page_number = request.GET.get(settings.MY_PAGE_NAME)
        projectList = paginator.get_page(page_number)

    else:
        projectList = None
        request.session['query_project'] = None
    return xFilter, projectList,isNotEmplyQuery



@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager', 'report-viewer'])
@manger_and_viewer_only
def manage_project(request, id=0):
    form_mode = 'NEW'
    if request.method == "GET":
        projectFilter, projectList,isNotEmplyQuery = filter_project(request)
        qstr = create_qstring_without_page(request)
        request.session['company_query'] = request.GET.get('company')

        if id == 0:  # new
            form = ProjectForm()
            form.fields["company"].queryset = Company.objects.filter(manager__user=request.user, is_customer=True)
            form_mode = 'NEW'
        else:  # edit
            project = get_object_or_404(Project, pk=id)
            form = ProjectForm(instance=project)
            form.fields["company"].queryset = Company.objects.filter(manager__user=request.user, is_customer=True)
            form.fields["company"].value = project.company
            form_mode = 'UPDATE'


    else:  # post data both adding and updating
        if id == 0:  # save from  new
            form_mode = 'NEW'
            form = ProjectForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, f'Project has been created successfully.')
            else:
                messages.error(request, form.errors)

        else:  # save from  edit
            form_mode = 'UPDATE'
            project = get_object_or_404(Project, pk=id)
            form = ProjectForm(request.POST, instance=project)
            if form.is_valid():
                form.save()
                messages.success(request, f'Project has been updated successfully.')
            else:
                messages.error(request, form.errors)

        # init new form after submiting
        return redirect('manage_project')

    return render(request, "app/project_manage.html",
                  {'form': form, 'projectFilter': projectFilter, 'projectList': projectList, 'mode': form_mode,
                   "qstr": qstr,"isNotEmplyQuery":isNotEmplyQuery
                   })



@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager', 'report-viewer'])
@manger_and_viewer_only
def manage_inventory(request):
    myuser = request.user
    qstr = create_qstring_without_page(request)
    all_qstr = create_all_qstring(request)
    isNotEmplyQuery = checkEmpyQueryString(request)
    request.session['company_query'] = request.GET.get('company')

    include_expired_inventory = None
    if isNotEmplyQuery:
        # recommnd filter by user and cust/yit/product warrant-end is less than today+- one month
        include_expired_inventory = request.GET.get('expired_inventory')
        if include_expired_inventory is None:
            my_all_inventories = Inventory.objects.filter(project__company__manager__user=myuser,
                                                          customer_warranty_end__gte=datetime.date.today()).order_by(
                '-id')
        else:
            my_all_inventories = Inventory.objects.filter(project__company__manager__user=myuser).order_by('-id')

    else:
        my_all_inventories = Inventory.objects.filter(project__company__manager__user=myuser,
                                                      customer_warranty_end__gte=datetime.date.today()).order_by(
            '-updated_at')[:settings.MY_PAGE_PER]

    inventoryFilter = InventoryFilter(request.GET, request=request, queryset=my_all_inventories)
    if inventoryFilter.qs.count() > 0:
        inventoryList = inventoryFilter.qs
        listIDs = [x.id for x in inventoryList]

        request.session['project_inventory'] = {"company": inventoryList[0].project.company.company_name}
        request.session['query_inventory'] = listIDs

        paginator = Paginator(inventoryList, settings.MY_PAGE_PER)
        page_number = request.GET.get(settings.MY_PAGE_NAME)
        inventoryList = paginator.get_page(page_number)
    else:
        inventoryList = None
        request.session['query_inventory'] = None
        request.session['project_inventory'] = None


    current_url_escaped = quote(request.get_full_path())

    context = {'inventoryList': inventoryList, 'inventoryFilter': inventoryFilter, "qstr": qstr, "all_qstr": all_qstr,
               'isNotEmplyQuery': isNotEmplyQuery, "current_url_escaped": current_url_escaped
               }
    return render(request, 'app/inventory_manage.html', context)


@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager', 'report-viewer'])
@manger_and_viewer_only
def manage_incident(request):
    myuser = request.user
    qstr = create_qstring_without_page(request)
    isNotEmplyQuery = checkEmpyQueryString(request)

    request.session['company_query'] = request.GET.get('company')
    if isNotEmplyQuery:
        my_all_incidents = Incident.objects.filter(inventory__project__company__manager__user=myuser)
    else:
        my_all_incidents = Incident.objects.filter(inventory__project__company__manager__user=myuser).order_by(
            '-incident_datetime')[:settings.MY_PAGE_PER]
    xFilter = IncidentFilter(request.GET, request=request, queryset=my_all_incidents)

    if xFilter.qs.count() > 0:
        incidentList = xFilter.qs
        listIDs = [x.id for x in incidentList]
        request.session['query_incident'] = listIDs
        request.session['project_incident'] = {"company": incidentList[0].inventory.project.company.company_name}

        paginator = Paginator(incidentList, settings.MY_PAGE_PER)
        page_number = request.GET.get(settings.MY_PAGE_NAME)
        incidentList = paginator.get_page(page_number)

    else:
        incidentList = None
        request.session['query_incident'] = None
        request.session['project_incident'] = None

    current_url_escaped = quote(request.get_full_path())

    date_from = request.GET.get('incident_datetime__gt')
    date_to = request.GET.get('incident_datetime__lt')
    date_between_qstring = ''
    if (date_from is not None) and (date_to is not None):
        date_between_qstring = f"?date_from={date_from}&date_to={date_to}"

    context = {'incidentList': incidentList, 'incidentFilter': xFilter, 'qstr': qstr,
               'isNotEmplyQuery': isNotEmplyQuery, 'current_url_escaped': current_url_escaped,
               'date_between_qstring': date_between_qstring}
    return render(request, 'app/incident_manage.html', context)


@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager', 'report-viewer'])
@manger_only
def delete_project(request, id):
    project_obj = Project.objects.get(pk=id)

    if request.method == "GET":
        inventory_list = project_obj.inventory_set.all()
        total_incident = sum([x.total_incidents() for x in inventory_list])

        pm_list = project_obj.preventivemaintenance_set.all()
        total_pm=pm_list.count()

    try:
        if request.method == "POST":
            project_obj.delete()

    except Exception as ex:
        messages.ERROR(ex)

    if request.method == "GET":
        context = {'project': project_obj, 'inventoryList': inventory_list, 'total_incident': total_incident,
                   'pmList': pm_list,'total_pm':total_pm}
        return render(request, 'app/project_delete.html', context)
    else:
        return redirect('manage_project')


def init_list_for_dropdownlist_inventory_form(form, branch, datacenter, list_customer_support, list_customer_pm_support,
                                              list_product_support,
                                              x_team, brand=None, model=None):
    form.fields['datacenter'].queryset = datacenter
    form.fields['branch'].queryset = branch

    form.fields['customer_support'].queryset = list_customer_support
    form.fields['customer_pm_support'].queryset = list_customer_pm_support

    form.fields['product_support'].queryset = list_product_support

    form.fields['cm_serviceteam'].queryset = x_team
    form.fields['pm_serviceteam'].queryset = x_team

    if brand is not None and model is not None:
        form.fields['brand'].queryset = brand
        form.fields['model'].queryset = model


def reload_prev_inventory_for_next_one(inventory_obj):
    # Error because Django is aware of the existing object  even though it want to set it for new one.
    # prev_form = InventoryForm(instance=inventory_obj)

    # The first row
    project_ref = get_object_or_404(Project, pk=inventory_obj.project.id) if inventory_obj.project is not None else None
    # serial_number_ref=inventory_obj.serial_number
    product_type_ref = get_object_or_404(Product_Type,
                                         pk=inventory_obj.product_type.id) if inventory_obj.product_type is not None else None
    brand_ref = get_object_or_404(Brand, pk=inventory_obj.brand.id) if inventory_obj.brand is not None else None
    model_ref = get_object_or_404(Model, pk=inventory_obj.model.id) if inventory_obj.model is not None else None

    template_ref=get_object_or_404(PM_Inventory_Template,pk=inventory_obj.pm_inventory_template.id) if inventory_obj.pm_inventory_template is not None else None

    customer_support_ref = get_object_or_404(Customer,
                                             pk=inventory_obj.customer_support.id) if inventory_obj.customer_support is not None else None
    customer_pm_support_ref = get_object_or_404(Customer,
                                                pk=inventory_obj.customer_pm_support.id) if inventory_obj.customer_pm_support is not None else None

    # The second row
    datacenter_ref = get_object_or_404(DataCenter,
                                       pk=inventory_obj.datacenter.id) if inventory_obj.datacenter is not None else None
    branch_ref = get_object_or_404(Branch, pk=inventory_obj.branch.id) if inventory_obj.branch is not None else None

    product_support_ref = get_object_or_404(Product,
                                            pk=inventory_obj.product_support.id) if inventory_obj.product_support is not None else None

    cm_serviceteam_ref = get_object_or_404(ServiceTeam,
                                           pk=inventory_obj.cm_serviceteam.id) if inventory_obj.cm_serviceteam is not None else None
    pm_serviceteam_ref = get_object_or_404(ServiceTeam,
                                           pk=inventory_obj.pm_serviceteam.id) if inventory_obj.pm_serviceteam is not None else None

    # The third row
    ref_customer_warranty_start = inventory_obj.customer_warranty_start
    ref_customer_warranty_end = inventory_obj.customer_warranty_end
    ref_cust_sla = get_object_or_404(SLA,
                                     pk=inventory_obj.customer_sla.id) if inventory_obj.customer_sla is not None else None

    ref_yit_warranty_start = inventory_obj.yit_warranty_start
    ref_yit_warranty_end = inventory_obj.yit_warranty_end
    ref_yit_sla = get_object_or_404(SLA, pk=inventory_obj.yit_sla.id) if inventory_obj.yit_sla is not None else None

    ref_product_warranty_start = inventory_obj.product_warranty_start
    ref_product_warranty_end = inventory_obj.product_warranty_end
    ref_product_sla = get_object_or_404(SLA,
                                        pk=inventory_obj.product_sla.id) if inventory_obj.product_sla is not None else None

    prev_form = InventoryForm(
        initial={
            'project': project_ref,
            'serial_number': '',
            'product_type': product_type_ref,
            'brand': brand_ref,
            'model': model_ref,
            'customer_support': customer_support_ref,
            'customer_pm_support': customer_pm_support_ref,

            'quantity': 1,
            'datacenter': datacenter_ref,
            'branch': branch_ref,
            'product_support': product_support_ref,
            'cm_serviceteam': cm_serviceteam_ref,
            'pm_serviceteam': pm_serviceteam_ref,

            'customer_warranty_start': ref_customer_warranty_start,
            'customer_warranty_end': ref_customer_warranty_end,
            'customer_sla': ref_cust_sla,

            'yit_warranty_start': ref_yit_warranty_start,
            'yit_warranty_end': ref_yit_warranty_end,
            'yit_sla': ref_cust_sla,

            'product_warranty_start': ref_product_warranty_start,
            'product_warranty_end': ref_product_warranty_end,
            'product_sla': ref_cust_sla,

            'pm_inventory_template':template_ref,

        }
    )

    return prev_form


@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager', 'report-viewer'])
@manger_and_viewer_only
def add_inventory(request, proj_id):

    isAccesible=app.user_access.check_user_to_do(request,proj_id,"AddInventory")
    if isAccesible==False:
        return HttpResponse(f'<h2>{request.user} are not authorized to manage any items of this company.</h2><b>Contact '
                            f'administrator to add this user to the manager OR engineer table in company</b>')


    project_obj = get_object_or_404(Project, pk=proj_id)
    list_inventory = Inventory.objects.filter(project_id=proj_id).order_by('-id')

    # product type brand , model , all item   is loaded because of no additional filter
    # data center ,branch and the remaining , they are loaded as dropdown list  based on some criteria like company and active status

    datacenter = DataCenter.objects.filter(company_id=project_obj.company_id)
    branch = Branch.objects.filter(company_id=project_obj.company_id)

    # PM and CM
    x_team = ServiceTeam.objects.filter(company__is_subcontractor=True)

    list_customer_support = Customer.objects.filter(company_id=project_obj.company_id, is_active=True)
    list_customer_pm_support = Customer.objects.filter(company_id=project_obj.company_id, is_active=True)
    list_product_support = Product.objects.filter(customer_company_id=project_obj.company_id, is_active=True)
    #list_template=PM_Inventory_Template.objects.all()

    if (request.method == "GET"):
        if len(list_inventory) == 0:
            form = InventoryForm(initial={'project': project_obj})
            init_list_for_dropdownlist_inventory_form(form, branch, datacenter, list_customer_support,
                                                      list_customer_pm_support,
                                                      list_product_support, x_team)

        else:
            inventory_prev_obj = list_inventory[0]
            form = reload_prev_inventory_for_next_one(inventory_prev_obj)

            brand_obj, model_obj = get_brand_its_model(inventory_prev_obj)
            init_list_for_dropdownlist_inventory_form(form, branch, datacenter, list_customer_support,
                                                      list_customer_pm_support,
                                                      list_product_support, x_team, brand_obj, model_obj)




    else:  # Post Data after submit

        form = InventoryForm(request.POST)
        # set list metat data for addid the next new one
        init_list_for_dropdownlist_inventory_form(form, branch, datacenter, list_customer_support,
                                                  list_customer_pm_support,
                                                  list_product_support, x_team)

        if form.is_valid():

            # save including its iteration Form to database
            inventory_obj = form.save()

            # New Version 26/4/22 , don't need to set any value except mode
            form.fields['model'].queryset = Model.objects.filter(brand_id=inventory_obj.brand_id, is_active=True)
            form.fields['model'].value = inventory_obj.model
            messages.success(request, 'Inventory has bee created successfully.')
        else:

            messages.error(request, form.errors)

    context = {'form': form, 'project': project_obj, 'inventoryList': list_inventory, 'mode': 'new'}
    return render(request, 'app/inventory_add_new.html', context)


@login_required(login_url='login')

@manger_and_viewer_only
def update_inventory(request, id):

    msg_check=app.user_access.check_user_to_do(request,id,"UpdateInventory")
    if msg_check is not None: return HttpResponse(msg_check)

    inventory_obj = get_object_or_404(Inventory, pk=id)
    project_obj = inventory_obj.project

    # product type brand , model , all item   is loaded because of no additional filter
    # data center ,branch and the remaining , they are loaded  as dropdown list based on some criteria like company and active status

    datacenter = DataCenter.objects.filter(company_id=inventory_obj.project.company_id)
    branch = Branch.objects.filter(company_id=inventory_obj.project.company_id)
    x_team = ServiceTeam.objects.filter(company__is_subcontractor=True)
    list_customer_support = Customer.objects.filter(company_id=project_obj.company_id, is_active=True)
    list_customer_pm_support = Customer.objects.filter(company_id=project_obj.company_id, is_active=True)

    # join customer who is inactive along wiht all acive customer
    if (inventory_obj.customer_support is not None) and (inventory_obj.customer_support.is_active == False):
        the_customer_support = Customer.objects.filter(id=inventory_obj.customer_support.id)
        list_customer_support = list_customer_support | the_customer_support

    if (inventory_obj.customer_pm_support is not None) and (inventory_obj.customer_pm_support.is_active == False):
        the_customer_pm_support = Customer.objects.filter(id=inventory_obj.customer_pm_support.id)
        list_customer_pm_support = list_customer_pm_support | the_customer_pm_support

    list_product_support = Product.objects.filter(customer_company_id=project_obj.company_id, is_active=True)
    if (inventory_obj.product_support is not None) and (inventory_obj.product_support.is_active == False):
        the_product_support = Product.objects.filter(id=inventory_obj.product_support_id)
        list_product_support = list_product_support | the_product_support

    if (request.method == "GET"):
        form = InventoryForm(instance=inventory_obj)

        brand_obj, model_obj = get_brand_its_model(inventory_obj)

        init_list_for_dropdownlist_inventory_form(form, branch, datacenter, list_customer_support,
                                                  list_customer_pm_support,
                                                  list_product_support, x_team, brand_obj, model_obj)
    else:
        form = InventoryForm(request.POST, instance=inventory_obj)
        init_list_for_dropdownlist_inventory_form(form, branch, datacenter, list_customer_support,
                                                  list_customer_pm_support,
                                                  list_product_support, x_team)
        if form.is_valid():
            if form.has_changed():
                form.save()
                next = request.GET.get('next', '/')

                if next == 'close_page':
                    return HttpResponse(
                        '<script type="text/javascript">window.close(); window.parent.location.href = "/";</script>')
                else:
                    return redirect(next)
            else:
                next = request.GET.get('next', '/')

                if next == 'close_page':
                    return HttpResponse(
                        '<script type="text/javascript">window.close(); window.parent.location.href = "/";</script>')
                else:
                    return redirect(next)



        else:
            messages.error(request, form.errors)

            # for view only

    context = {}
    if request.GET.get('view_only'):
        context = {'form': form, 'project': project_obj, 'xview': 1, 'mode': 'update'}
    else:
        context = {'form': form, 'project': project_obj, 'xview': None, 'mode': 'update'}
    return render(request, 'app/inventory_update.html', context)


def get_brand_its_model(inventory_obj):
    if inventory_obj.model.is_active == True:
        brand_obj = Brand.objects.all()
        model_obj = Model.objects.filter(brand_id=inventory_obj.brand_id, is_active=True)
        is_active = True
    else:
        brand_obj = Brand.objects.filter(id=inventory_obj.brand_id)
        model_obj = Model.objects.filter(id=inventory_obj.model.id)
        is_active = False
    return brand_obj, model_obj


@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager', 'report-viewer'])
@manger_only
def copy_inventoryList_by_existingProject(request, proj_ref_id):
    projectRef_obj = get_object_or_404(Project, pk=proj_ref_id)
    inventoryList_copy = Inventory.objects.filter(project_id=proj_ref_id)

    # get query string  to check wheather warranty_date_once
    onetime_warranty_date = request.GET.get('onetime_warranty_date')
    if onetime_warranty_date is not None:
        onetime_warranty_date = int(onetime_warranty_date)
        if onetime_warranty_date not in [0, 1]:
            return redirect('manage_project')
    else:
        onetime_warranty_date = -1

    if request.method == 'GET':

        project_form = ProjectForm(initial={'company': projectRef_obj.company})
        project_form.fields["company"].queryset = Company.objects.filter(id=projectRef_obj.company_id)

        inventory_formset = InventoryFormset(queryset=inventoryList_copy)
        for invt_form in inventory_formset:
            invt_id = invt_form.fields['id'].initial
            x_inventory = Inventory.objects.get(id=invt_id)

            invt_form.fields['brand'].queryset = Brand.objects.filter(id=x_inventory.brand_id)
            invt_form.fields['brand'].value = x_inventory.brand
            invt_form.fields['model'].queryset = Model.objects.filter(id=x_inventory.model_id)
            invt_form.fields['model'].value = x_inventory.model

            invt_form.fields['product_type'].queryset = Product_Type.objects.filter(id=x_inventory.product_type_id)
            invt_form.fields['product_type'].value = x_inventory.product_type

    elif request.method == 'POST':
        # project_form.fields["company"].value=projectRef_obj.company
        project_form = ProjectForm(request.POST)
        inventory_formset = InventoryFormset(request.POST)

        if project_form.is_valid() and inventory_formset.is_valid():

            project_obj = project_form.save()

            qty = 0
            cust_start_first = None
            cust_end_first = None
            cust_sla_first = None
            yit_start_first = None
            yit_end_first = None
            yit_sla_first = None
            product_start_first = None
            product_end_first = None
            product_sla_first = None

            count_item = 1
            for inventory_form in inventory_formset:
                if inventory_form not in inventory_formset.deleted_forms:

                    inventory_obj = inventory_form.save(commit=False)
                    inventory_obj.id = None  # create new inventory from existing one (Coppy)
                    inventory_obj.project = project_obj

                    if onetime_warranty_date == 1 and count_item == 1:
                        cust_start_first = inventory_obj.customer_warranty_start
                        cust_end_first = inventory_obj.customer_warranty_end
                        cust_sla_first = inventory_obj.customer_sla
                        yit_start_first = inventory_obj.yit_warranty_start
                        yit_end_first = inventory_obj.yit_warranty_end
                        yit_sla_first = inventory_obj.yit_sla
                        product_start_first = inventory_obj.product_warranty_start
                        product_end_first = inventory_obj.product_warranty_end
                        product_sla_first = inventory_obj.product_sla
                    elif onetime_warranty_date == 1 and count_item > 1:
                        inventory_obj.customer_warranty_start = cust_start_first
                        inventory_obj.customer_warranty_end = cust_end_first
                        inventory_obj.customer_sla = cust_sla_first
                        inventory_obj.yit_warranty_start = yit_start_first
                        inventory_obj.yit_warranty_end = yit_end_first
                        inventory_obj.yit_sla = yit_sla_first
                        inventory_obj.product_warranty_start = product_start_first
                        inventory_obj.product_warranty_end = product_end_first
                        inventory_obj.product_sla = product_sla_first

                    inventory_obj.save()
                    inventory_id = inventory_obj.id
                    count_item += 1

            return redirect('add_inventory', proj_id=project_obj.id)
            # return redirect('manage_project')

        else:
            messages.error(request, project_form.errors)
            messages.error(request, inventory_formset.errors)

    context = {'project_form': project_form, 'inventory_formset': inventory_formset, 'project_ref': projectRef_obj,
               'inventory_count': inventoryList_copy.count(), "onetime_warranty_date": onetime_warranty_date
               }
    return render(request, 'app/inventory_add_copy.html', context)


@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager', 'report-viewer'])
@manger_only
def delete_inventory(request, id):

    msg_check=app.user_access.check_user_to_do(request,id,"DeleteInventory")
    if msg_check is not None: return HttpResponse(msg_check)

    inventory_obj = Inventory.objects.get(pk=id)

    if request.method == "GET":
        incident_list = inventory_obj.incident_set.all()
        total_incident = incident_list.count()

        total_pmITem= inventory_obj.pm_inventory_set.filter(is_pm=True).count()


    try:
        if request.method == "POST":
            inventory_obj.delete()

    except Exception as ex:
        messages.ERROR(ex)

    if request.method == "GET":
        context = {'inventory': inventory_obj,
                   'incidentList': incident_list, 
                   'total_incident': total_incident,'total_pm_item': total_pmITem
                   }
        return render(request, 'app/inventory_delete.html', context)
    else:
        return redirect('manage_inventory')

@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager', 'report-viewer'])
@manger_only

def add_incident(request, inventory_id):

    msg_check=app.user_access.check_user_to_do(request,inventory_id,"AddIncident")
    if msg_check is not None: return HttpResponse(msg_check)

    inventory_obj = get_object_or_404(Inventory, pk=inventory_id)

    if (request.method == "GET"):
        incident_form = IncidentForm(initial={'incident_status': 1})
        incident_form.fields['incident_owner'].queryset =Employee.objects.filter(is_inactive=False)

        file_form = IncidentFileForm()

        incidentDetail_formset = Incident_DetailFormset(queryset=Incident_Detail.objects.none())

    elif request.method == 'POST':

        incident_form = IncidentForm(request.POST)
        file_form = IncidentFileForm(request.POST, request.FILES)
        incidentDetail_formset = Incident_DetailFormset(request.POST)

        if incident_form.is_valid():
            if file_form.is_valid():
                if incidentDetail_formset.is_valid():

                    # save incident
                    incident_obj = incident_form.save(commit=False)

                    incident_obj.inventory = inventory_obj
                    incident_obj.save()

                    year_code = int(datetime.date.today().strftime('%y'))
                    id_str = str(incident_obj.id)
                    id_str = id_str.zfill(4)
                    incident_code = f'SR-ES-{year_code}-{id_str}'
                    incident_obj.incident_no = incident_code
                    incident_obj.save()

                    # upload file
                    files = upload_incident_files(request, incident_obj)

                    # save incident detai
                    for form in incidentDetail_formset:
                        x = form.cleaned_data
                        if x != {}:
                            detail_obj = form.save(commit=False)

                            detail_obj.incident_master = incident_obj
                            detail_obj.save()

                    # return http://127.0.0.1:8000/update_incident/21/
                    return redirect('update_incident', id=incident_obj.id)
                    # return redirect('/incidents')
                else:
                    messages.error(request, incidentDetail_formset.errors)
            else:
                messages.error(request, file_form.errors)
        else:
            messages.error(request, incident_form.errors)

    incident_basic_info = {'id': '',
                           'incident_no': '',
                           'serial_number': inventory_obj.serial_number,
                           'enq_id': inventory_obj.project.enq_id,
                           'company': inventory_obj.project.company.company_name,
                           'project_name': inventory_obj.project.project_name
                           }

    context = {'inventory': inventory_obj, 'incident_basic_info': incident_basic_info,
               'FILE_MAX_SIZE_MB': settings.UPLOAD_FILE_MAX_SIZE_MB,
               'incident_form': incident_form, 'file_form': file_form, 'formset': incidentDetail_formset}
    return render(request, 'app/incident_add.html', context)


@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager', 'report-viewer'])
@manger_and_viewer_only
def update_incident(request, id):

    msg_check=app.user_access.check_user_to_do(request,id,"UpdateIncident")
    if msg_check is not None: return HttpResponse(msg_check)

    predSeverity = ''
    biPredSeverity=''

    incident_obj = get_object_or_404(Incident, pk=id)
    inventory_obj = incident_obj.inventory;

    incident_detailList = incident_obj.incident_detail_set.all()
    incident_files = incident_obj.files.all()

    # =======================load filter for change inventory only same comanpy as dummy==============================================
    myuser = request.user
    isNotEmplyQuery = checkEmpyQueryString(request)
    if isNotEmplyQuery:
        my_all_inventories = Inventory.objects.filter(project__company__manager__user=myuser,
                                                      project__company__id=inventory_obj.project.company_id,
                                                      ).exclude(id=inventory_obj.id)
    else:
        my_all_inventories = Inventory.objects.filter(id=0)

    inventoryFilter = InventoryToChangeInIncidentFilter(request.GET, queryset=my_all_inventories)

    if inventoryFilter.qs.count() > 0:
        inventoryList = inventoryFilter.qs
    else:
        inventoryList = None
    # =====================================================================

    if (request.method == "GET"):

        # ===============load ml prediction====================
        if incident_obj.incident_status.id==settings.INCIDENT_CODE_CLOSED:

         incidentPredictions =incident_obj.prediction_ml_severity_incident_set.all()
         if len(incidentPredictions)>0:
             predObj=incidentPredictions.latest('imported_at')
             predSeverity=f"Severity-Prediction = {predObj.severity_name}"
         binaryPrediction=incident_obj.prediction_ml2_everity_incident_set.all()
         if len(binaryPrediction)>0:
             biPredObj=binaryPrediction.latest('imported_at')
             biPredSeverity = f" (Level:{biPredObj.severity_name})"

         #===============================================================


        # ==============load  incident form and it detail=================
        incident_form = IncidentForm(instance=incident_obj)
        incident_form.fields['incident_owner'].queryset = Employee.objects.filter(
            Q(is_inactive=False) | Q(id = incident_obj.incident_owner.id))
        # is_inactive=False
        file_form = IncidentFileForm(initial={'incident_ref': incident_obj})

    else:  # post
        incident_form = IncidentForm(request.POST, instance=incident_obj)
        file_form = IncidentFileForm(request.POST, request.FILES)

        if (incident_form.is_valid()):
            if (file_form.is_valid()):

                status = incident_form.cleaned_data['incident_status']
                if status.id==settings.INCIDENT_CODE_CLOSED and incident_obj.incident_detail_set.count()==0:
                    messages.error(request, "You are not allowed to close an incident because of no incident detail.")
                # update incident
                else:
                    if incident_form.has_changed():
                        incident_form.save()
                        messages.success(request, f'Incident has been updated successfully.')
                    else:
                        messages.success(request, f'No any update in this incident.')

                    # upload file
                    files = upload_incident_files(request, incident_obj)
                    if (len(files) > 0):
                        incident_files = Incident_File.objects.filter(incident_ref_id=id)
            else:
                messages.success(request, file_form.errors)

        else:
            messages.error(request, incident_form.errors)

    incident_basic_info = {'id': incident_obj.id,
                           'incident_no': incident_obj.incident_no,
                           'serial_number': incident_obj.inventory.serial_number,
                           'enq_id': incident_obj.inventory.project.enq_id,
                           'company': incident_obj.inventory.project.company.company_name,
                           'project_name': incident_obj.inventory.project.project_name

                           }

    context = {'inventory': inventory_obj, 'incident_form': incident_form,
               'incident_basic_info': incident_basic_info, 'FILE_MAX_SIZE_MB': settings.UPLOAD_FILE_MAX_SIZE_MB,
               'incident_detailList': incident_detailList,
               'incident_files': incident_files, 'file_form': file_form,
               'inventoryFilter': inventoryFilter, 'inventoryList': inventoryList
               ,'predSeverity':predSeverity,'biPredSeverity':biPredSeverity
               }
    return render(request, 'app/incident_update.html', context)


@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager', 'report-viewer'])
@manger_only
def change_inventory_for_incident(request, id, inventory_id):
    msg_check=app.user_access.check_user_to_do(request,inventory_id,"ChangeInventory")
    if msg_check is not None: return HttpResponse(msg_check)


    incident_obj = get_object_or_404(Incident, pk=id)
    inventory_obj = get_object_or_404(Inventory, pk=inventory_id)
    try:
        if request.method == "GET":
            incident_obj.inventory = inventory_obj
            incident_obj.save()

            next = request.GET.get('next', '/')
            return redirect('update_incident', incident_obj.id)
    except Exception as ex:
        messages.ERROR(ex)


@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager', 'report-viewer'])
@manger_only
def delete_incident(request, id):

    msg_check=app.user_access.check_user_to_do(request,id,"DeleteIncident")
    if msg_check is not None: return HttpResponse(msg_check)

    x_obj = Incident.objects.get(pk=id)

    if request.method == "GET":
        detail_list = x_obj.incident_detail_set.all()
        total_detail = detail_list.count()
        is_closed_status = 1 if x_obj.incident_status.id == settings.INCIDENT_CODE_CLOSED else 0

    try:
        if request.method == "POST":
            x_obj.delete()

    except Exception as ex:
        messages.ERROR(ex)
    if request.method == "GET":
        context = {'incident': x_obj, 'detailList': detail_list,
                   'total_detail': total_detail, 'is_closed_status': is_closed_status}
        return render(request, 'app/incident_delete.html', context)
    else:
        return redirect('manage_incident')


def filter_detail(request, incident_id):
    myuser = request.user
    my_all = Incident_Detail.objects.filter(incident_master_id=incident_id)

    listEngineerByIncidentDetail=list( my_all.values_list('employee_id',flat=True).distinct())
    request.session['listEngineerByIncidentDetail']=listEngineerByIncidentDetail

    xFilter = DetailFilter(request.GET, request=request, queryset=my_all)
    xList = xFilter.qs

    return xFilter, xList


@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager', 'report-viewer'])
@manger_and_viewer_only
def manage_incident_detail(request, incident_id, id=0):

    msg_check=app.user_access.check_user_to_do(request,incident_id,"UpdateDetail")
    if msg_check is not None: return HttpResponse(msg_check)

    detailFilter, incident_detailList = filter_detail(request, incident_id)
    incident_obj = get_object_or_404(Incident, pk=incident_id)

    code_closed = settings.INCIDENT_CODE_CLOSED
    incident_inprogress = True

    if incident_obj.incident_status.id == code_closed:
        incident_inprogress = False

    if request.method == "GET":
        if id == 0:  # new detail
            form = Incident_DetailForm()
            form_mode = 'new'
        else:  # edit
            detail = get_object_or_404(Incident_Detail, pk=id)
            form = Incident_DetailForm(instance=detail)
            form_mode = 'update'

    else:  # post data both adding and updating

        if id == 0:  # save from  new
            form = Incident_DetailForm(request.POST)
            form_mode = 'new'
            if form.is_valid():
                new_detail_obj = form.save(commit=False)
                new_detail_obj.incident_master = incident_obj
                new_detail_obj.save()
                messages.success(request, f'Detail has been created successfully.')
                return redirect('manage_incident_detail', incident_id, 0)
            else:
                messages.error(request, form.errors)
        else:  # save from  edit
            detail = get_object_or_404(Incident_Detail, pk=id)
            form = Incident_DetailForm(request.POST, instance=detail)
            form_mode = 'update'
            if form.is_valid():
                if form.has_changed():
                    detail_obj = form.save(commit=False)
                    detail_obj.incident_master = incident_obj
                    detail_obj.save()
                    messages.success(request, f'Detail has been updated successfully.')
                else:
                    messages.info(request, f'No any update.')

                return redirect('manage_incident_detail', incident_id, 0)
            else:
                messages.error(request, form.errors)

    context = {'incident_id': incident_id, 'form': form, 'incident_detailList': incident_detailList, 'mode': form_mode,
               'detailFilter': detailFilter, 'incident_inprogress': incident_inprogress}
    return render(request, "app/incident_detail_manage.html", context)


@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager', 'report-viewer'])
@manger_only
def delete_incident_detail(request, id):
    msg_check=app.user_access.check_user_to_do(request,id,"DeleteDetail")
    if msg_check is not None: return HttpResponse(msg_check)

    obj = Incident_Detail.objects.get(pk=id)
    master_id = obj.incident_master_id
    try:
        if request.method == "GET":
            obj.delete()
            return redirect('manage_incident_detail', master_id, 0)
    except Exception as ex:
        messages.ERROR(ex)


@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager', 'report-viewer'])
@manger_and_viewer_only
def delete_incident_file(request, id):
    file = Incident_File.objects.get(pk=id)
    try:
        if request.method == "GET":
            file.delete()
            # delete file
            next = request.GET.get('next', '/')
            return redirect(next)
    except Exception as ex:
        messages.ERROR(ex)


def load_models_by_brand(request):
    brand_idx = request.GET.get('brand')
    models = Model.objects.filter(brand_id=brand_idx, is_active=True).order_by('model_name')

    return render(request, 'app/models_by_brand_list.html', {'models': models})


def load_customers_support_by_company(request):
    idx = request.GET.get('company')
    list_customer_support = Customer.objects.filter(company_id=idx, is_active=True).order_by('customer_name')
    return render(request, 'app/customers_by_company_list.html', {'list_customer_support': list_customer_support})


@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager'])
# @staff_admin_only
@manger_only
# ccreate an excel template to ease  site manager  to fill in   model data  this form
def create_model_template(request):
    if request.method == "POST":

        # create excel as specified template
        brandList_form = ExportBrandForm(request.POST)
        if (brandList_form.is_valid()):

            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output)
            bold = workbook.add_format({'bold': True})

            brandList = brandList_form.cleaned_data.get('brand_list')

            for brand in brandList:
                worksheet = workbook.add_worksheet(name=brand.brand_name)
                # worksheet.add_table('A1:A21')
                # worksheet.set_column('A1:A21', 20)
                worksheet.write('A1', 'model_name', bold)
                worksheet.write('B1', 'model_description', bold)

            workbook.close()

            output.seek(0)

            filename = 'import_models_template.xlsx'
            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=%s' % filename

            return response
    else:
        brandList_form = ExportBrandForm()
        modelImport_form = UploadModelForm()
    context = {'brandList_form': brandList_form, 'modelImport_form': modelImport_form}
    return render(request, 'app/import_models.html', context)


@login_required(login_url='login')
@manger_only
def upload_models_template(request):
    save_button = False
    list_model = []
    list_model_json = []
    brandList_form = ExportBrandForm()
    modelImport_form = UploadModelForm()

    if request.method == "POST":

        form = UploadModelForm(request.POST, request.FILES)

        if form.is_valid():
            file = form.cleaned_data['template_file']

            wb = openpyxl.load_workbook(file)
            for sheet in wb.worksheets:
                try:
                    x_brand = re.sub('\s+', ' ', sheet.title.strip())
                    brand_obj = Brand.objects.get(brand_name__iexact=x_brand)

                    b_id = brand_obj.id
                    b_name = brand_obj.brand_name

                except ObjectDoesNotExist as ex:
                    b_id = 0
                    b_name = sheet.title
                    messages.info(request, ex)

                for row in sheet.iter_rows(min_row=2, max_col=2):
                    # for i,row_cells in enumerate(sheet.iter_rows(values_only=True)):
                    try:

                        model_obj = ImportedModel(brand_id=0, brand_name='', model_name='', note='', isValid=True,
                                                  model_des='')

                        x_model = re.sub('\s+', ' ', row[0].value.strip())
                        x_des = row[1].value
                        model_obj.model_name = x_model
                        model_obj.model_des = x_des

                        if model_obj.model_des == '' or model_obj.model_des is None:
                            model_obj.model_des = '-'

                        # count_exst_models = Model.objects.filter(model_name__iexact=model_obj.model_name, brand__brand_name__iexact=x_brand).count()
                        count_exst_models = Model.objects.filter(model_name__iexact=model_obj.model_name).count()

                        model_obj.brand_id = b_id
                        model_obj.brand_name = b_name

                        if model_obj.brand_id == 0:
                            model_obj.isValid = False
                            model_obj.note = f'{model_obj.note}Invalid Brand , '

                        if count_exst_models > 0:
                            model_obj.isValid = False
                            model_obj.note = f'{model_obj.note}Existing Model in database(Cannot be inserted the item into database),'

                        count_dup = [x for x in list_model if
                                     (x.model_name.lower() == model_obj.model_name.lower()) and (
                                             x.brand_name.lower() == model_obj.brand_name.lower())]
                        if len(count_dup) > 0:
                            model_obj.isValid = False
                            model_obj.note = f'{model_obj.note}Deplicated Model in excel (Just be inserted new one),'

                        list_model.append(model_obj)
                        list_model_json.append(model_obj.to_json())

                    except Exception as ex:
                        pref_error = f'{x_brand} - {row[0]}'
                        messages.info(request, f'{pref_error} - {ex}')

    # list_model.sort(key=operator.attrgetter(['isValid','model_name']))
    list_model.sort(key=lambda x: (x.isValid, x.model_name))

    request.session['imported_models'] = list_model_json

    valid_some = [item for item in list_model if item.isValid == True]
    if (len(valid_some) > 0):
        save_button = True
    else:
        save_button = False

    return render(request, 'app/import_models.html',
                  {'list_model': list_model, 'brandList_form': brandList_form, 'modelImport_form': modelImport_form,
                   "save_button": save_button})


@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager'])
# @staff_admin_only
@manger_only
def import_models(request):
    brandList_form = ExportBrandForm()
    modelImport_form = UploadModelForm()

    if request.method == "POST":
        list_model_json = []
        list_model = request.session.get('imported_models', list_model_json)
        for obj in list_model:
            model_dict = json.loads(obj)
            x_valid = model_dict['isValid']
            if x_valid == True:
                name = model_dict['model_name']
                des = model_dict['model_des']

                x_brand = model_dict['brand_name']
                x_brand_id = model_dict['brand_id']
                Model.objects.create(brand_id=x_brand_id, model_name=name, model_description=des)

        return redirect('manage_project')

    return render(request, 'app/import_models.html',
                  {'brandList_form': brandList_form, 'modelImport_form': modelImport_form})


@login_required(login_url='login')
# @allowed_users(allowed_roles=['administrator'])
@manger_only
def manage_sm(request):
    return render(request, 'app/aadmin_manage_sitemanager.html')


def upload_incident_files(request, incident_obj):
    files = request.FILES.getlist('incident_file')
    for f in files:
        file_instance = Incident_File(incident_file=f, incident_ref=incident_obj)
        file_instance.save()
        # messages.success(request, f'Files {f.name} have been uploaded successfully.')
    return files



def export_data_refactor(request, query_session):

    listIDs = request.session.get(query_session, None)

    if listIDs != None:
        date_btw = ''
        buildtime = datetime.datetime.now().strftime('%d%m%y_%H%M')

        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        if date_from is not None and date_to is not None:
            date_btw = f'_From{date_from}_To{date_to}'

        company_query = request.session['company_query']
        if query_session == 'query_project' or query_session == 'query_all_project':
            df=reporter.report_project(listIDs)
            if df.empty:
                messages.error(request, "No project return")
                return render(request, 'app/export_report.html', {})
            if (company_query is not None) and (company_query != ''):
                comanpy_info=Company.objects.get(id=int(company_query ))
                file_name = f"Project_{comanpy_info.company_name}_{buildtime}.xlsx"
            else:
                file_name = f"Project_{buildtime}.xlsx"
            request.session['company_query'] = None

        elif query_session == 'query_inventory':
            df = reporter.report_inventory(listIDs)
            if df.empty:
                messages.error(request, "No inventory return except dummy inventory")
                return render(request, 'app/export_report.html', {})

            comanpy_info = request.session.get("project_inventory", None)
            if (comanpy_info is not None) and (company_query is not None) and (company_query != ''):
                file_name = f"Inventory_{comanpy_info['company']}{date_btw}_{buildtime}.xlsx"
            else:
                file_name = f"Inventory{date_btw}_{buildtime}.xlsx"
        # elif  query_session=='query_incident':
        else:  # query_incident
            df = reporter.report_incident(listIDs)
            if df.empty:
                messages.error(request, "No incident return")
                return render(request, 'app/export_report.html', {})
            comanpy_info = request.session.get("project_incident", None)
            if (comanpy_info is not None) and (company_query is not None) and (company_query != ''):
                file_name = f"Incident_{comanpy_info['company']}{date_btw}_{buildtime}.xlsx"
            else:
                file_name = f"Incident{date_btw}_{buildtime}.xlsx"

        df = df.fillna(value='')

        col_remove = 'ID'
        if col_remove in df.columns:
            df = df.drop([col_remove], axis=1)

        if len(df) > 0:
            try:
                wb = openpyxl.Workbook(write_only=True)
                ws = wb.create_sheet()
                for r in dataframe_to_rows(df, index=False, header=True, ):
                    ws.append(r)
            except openpyxl.utils.exceptions.IllegalCharacterError as e:
                raise e

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'


        wb.save(response)

        return response

        context = {}
        return render(request, 'app/export_report.html', context)


@login_required(login_url='login')
@manger_and_viewer_only
def export_project(request):
    response = export_data_refactor(request, 'query_project')
    return response

@login_required(login_url='login')
@manger_and_viewer_only
def export_all_project(request):
    myuser = request.user
    projectList = Project.objects.filter(company__manager__user=myuser, project_end__gte=datetime.date.today(),                             is_dummy=False).order_by('-id')
    if len(projectList)>0:
     listIDs = [x.id for x in projectList]
     request.session['query_all_project'] = listIDs
    else:
        request.session['query_all_project']=None
    response = export_data_refactor(request, 'query_all_project')
    return response

@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager', 'report-viewer'])
@manger_and_viewer_only
def export_inventory(request):
    response = export_data_refactor(request, 'query_inventory')
    return response


@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager', 'report-viewer'])
@manger_and_viewer_only
def export_incident(request):
    response = export_data_refactor(request, 'query_incident')
    return response


def filter_location(request, location_type):
    myuser = request.user
    isNotEmplyQuery = checkEmpyQueryString(request)
    if location_type == 1:  # branch
        if isNotEmplyQuery:
            my_all = Branch.objects.filter(company__manager__user=myuser).order_by('-updated_at')
        else:
            my_all = Branch.objects.filter(company__manager__user=myuser).order_by('-updated_at')[:settings.MY_PAGE_PER]
        xFilter = BranchFilter(request.GET, request=request, queryset=my_all)

    else:  # if location_type == 2  datacenter
        if isNotEmplyQuery:
            my_all = DataCenter.objects.filter(company__manager__user=myuser).order_by('-updated_at')
        else:
            my_all = DataCenter.objects.filter(company__manager__user=myuser).order_by('-updated_at')[
                     :settings.MY_PAGE_PER]

        xFilter = DataCenterFilter(request.GET, request=request, queryset=my_all)

    if xFilter.qs.count() > 0:
        xList = xFilter.qs

    else:
        xList = None
    return xFilter, xList


@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager'])
@manger_only
def manage_branch(request, id=0):
    myuser = request.user

    if request.method == "GET":
        branchFilter, branchList = filter_location(request, 1)
        if id == 0:  # new detail
            form = BranchForm()
            form.fields['company'].queryset = Company.objects.filter(manager__user=myuser, is_customer=True)
            form_mode = 'New'
        else:
            obj = get_object_or_404(Branch, pk=id)
            form = BranchForm(instance=obj)

            form.fields['company'].queryset = Company.objects.filter(manager__user=myuser, is_customer=True)
            form.fields['customer'].queryset = Customer.objects.filter(
                Q(company_id=obj.company_id, is_active=True) | Q(id=obj.customer_id))

            form_mode = 'Update'

    else:
        if id == 0:  # save from  new
            form = BranchForm(request.POST)
            form_mode = 'New'
            if form.is_valid():
                form.save()
                messages.success(request, f'Branch has been created successfully.')
            else:
                messages.error(request, form.errors)
        else:  # save from  edit
            obj = get_object_or_404(Branch, pk=id)
            form = BranchForm(request.POST, instance=obj)
            form_mode = 'Update'
            if form.is_valid():
                if form.has_changed():
                    form.save()
                    messages.success(request, f'Branch has been updated successfully.')
                else:
                    messages.info(request, f'No any update for Branch.')
            else:
                messages.error(request, form.errors)

        return redirect('manage_branch', id=0)

    html_template = "app/location_branch_manager.html"
    context = {'form': form, 'mode': form_mode, 'locationList': branchList, 'locationFilter': branchFilter}
    return render(request, html_template, context)


@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager'])
@manger_only
def manage_datacenter(request, id=0):
    myuser = request.user

    if request.method == "GET":
        dcFilter, dcList = filter_location(request, 2)
        if id == 0:  # new detail
            form = DataCenterForm()
            form.fields['company'].queryset = Company.objects.filter(manager__user=myuser, is_customer=True)
            form_mode = 'New'
        else:
            obj = get_object_or_404(DataCenter, pk=id)
            form = DataCenterForm(instance=obj)

            form.fields['company'].queryset = Company.objects.filter(manager__user=myuser, is_customer=True)
            form.fields['customer'].queryset = Customer.objects.filter(
                Q(company_id=obj.company_id, is_active=True) | Q(id=obj.customer_id))

            form_mode = 'Update'

    else:
        if id == 0:  # save from  new
            form = DataCenterForm(request.POST)
            form_mode = 'New'
            if form.is_valid():
                form.save()
                messages.success(request, f'DataCenter has been created successfully.')
            else:
                messages.error(request, form.errors)
        else:  # save from  edit
            obj = get_object_or_404(DataCenter, pk=id)
            form = DataCenterForm(request.POST, instance=obj)
            form_mode = 'Update'
            if form.is_valid():
                if form.has_changed():
                    form.save()
                    messages.success(request, f'DataCenter has been updated successfully.')
                else:
                    messages.info(request, f'No any update for DataCenter.')
            else:
                messages.error(request, form.errors)

        return redirect('manage_datacenter', id=0)

    html_template = "app/location_datacenter_manager.html"
    context = {'form': form, 'mode': form_mode, 'locationList': dcList, 'locationFilter': dcFilter}
    return render(request, html_template, context)


def filter_supporter(request, support_type):
    myuser = request.user
    isNotEmplyQuery = checkEmpyQueryString(request)
    if support_type == 1:
        if isNotEmplyQuery:
            my_all = Customer.objects.filter(company__manager__user=myuser).order_by('-updated_at')
        else:
            my_all = Customer.objects.filter(company__manager__user=myuser).order_by('-updated_at')[
                     :settings.MY_PAGE_PER]
        xFilter = CustomerSupportFilter(request.GET, request=request, queryset=my_all)

    else:
        if isNotEmplyQuery:
            my_all = Product.objects.filter(customer_company__manager__user=myuser).order_by('-updated_at')
        else:
            my_all = Product.objects.filter(customer_company__manager__user=myuser).order_by('-updated_at')[
                     :settings.MY_PAGE_PER]

        xFilter = ProductSupportFilter(request.GET, request=request, queryset=my_all)

    if xFilter.qs.count() > 0:
        xList = xFilter.qs

    else:
        xList = None
    return xFilter, xList


# 1= customer specified and 2=product_type
@login_required(login_url='login')
# @allowed_users(allowed_roles=['site-manager'])
@manger_only
def manage_supporter(request, support_type, id=0):
    myuser = request.user

    # personSupportList = Customer.objects.filter(company__manager__user=myuser ) if support_type==1  else Product.objects.filter(customer_company__manager__user=myuser)
    objectName = 'Customer Support' if support_type == 1 else 'Product Support'

    if request.method == "GET":
        personSupportFilter, personSupportList = filter_supporter(request, support_type)
        if id == 0:  # new detail
            if support_type == 1:
                form = CustomerForm()
                form.fields['company'].queryset = Company.objects.filter(manager__user=myuser, is_customer=True)

            else:
                form = ProductForm()
                form.fields['customer_company'].queryset = Company.objects.filter(manager__user=myuser,
                                                                                  is_customer=True)
                form.fields['partner_company'].queryset = Company.objects.filter(is_subcontractor=True)

            form_mode = 'New'

        else:  # edit
            if support_type == 1:
                obj = get_object_or_404(Customer, pk=id)
                form = CustomerForm(instance=obj)
                form.fields['company'].queryset = Company.objects.filter(manager__user=myuser, is_customer=True)
            else:
                obj = get_object_or_404(Product, pk=id)
                form = ProductForm(instance=obj)
                form.fields['customer_company'].queryset = Company.objects.filter(manager__user=myuser,
                                                                                  is_customer=True)
                form.fields['partner_company'].queryset = Company.objects.filter(is_subcontractor=True)

            form_mode = 'Update'

    else:  # post data both adding and updating
        if id == 0:  # save from  new
            form = CustomerForm(request.POST) if support_type == 1 else ProductForm(request.POST)
            form_mode = 'New'
            if form.is_valid():
                form.save()
                messages.success(request, f'{objectName} has been created successfully.')
            else:
                messages.error(request, form.errors)
        else:  # save from  edit
            obj = get_object_or_404(Customer, pk=id) if support_type == 1 else get_object_or_404(Product, pk=id)
            form = CustomerForm(request.POST, instance=obj) if support_type == 1 else ProductForm(request.POST,
                                                                                                  instance=obj)
            form_mode = 'Update'
            if form.is_valid():
                if form.has_changed():
                    form.save()

                    messages.success(request, f'{objectName} has been updated successfully.')
                else:
                    messages.info(request, f'No any update for {objectName}.')
            else:
                messages.error(request, form.errors)

        return redirect('manage_supporter', support_type=support_type, id=0)

    html_template = "app/supporter_customer.html" if support_type == 1 else "app/supporter_product.html"
    context = {'form': form, 'personSupportFilter': personSupportFilter, 'personSupportList': personSupportList,
               'mode': form_mode, 'object_name': objectName}
    return render(request, html_template, context)


def manage_xyz(request, id=0):
    xyzList = XYZ_TestData.objects.all()
    form_mode = 'new'
    if request.method == "GET":
        if id == 0:  # new detail
            form = XYZ_TestDataForm()
            form_mode = 'new'
        else:  # edit
            xyz = get_object_or_404(XYZ_TestData, pk=id)
            form = XYZ_TestDataForm(instance=xyz)
            form_mode = 'update'

    else:  # post data both adding and updating
        if id == 0:  # save from  new
            form_mode = 'new'
            form = XYZ_TestDataForm(request.POST)
        else:  # save from  edit
            form_mode = 'update'
            xyz = get_object_or_404(XYZ_TestData, pk=id)
            form = XYZ_TestDataForm(request.POST, instance=xyz)
        if form.is_valid():
            form.save()

        else:
            messages.error(request, form.errors)
        return redirect('manage_xyz')

    return render(request, "app/test_my_xyz.html",
                  {'form': form, 'xyzList': xyzList, 'mode': form_mode})


def contact(request):
    """Renders the contact page."""
    assert isinstance(request, HttpRequest)
    return render(request,
                  'app/contact.html',
                  {
                      'title': 'Contact',
                      'message': 'Your contact page.',
                      'year': datetime.now().year,
                  })


def about(request):
    """Renders the about page."""
    assert isinstance(request, HttpRequest)
    return render(request,
                  'app/about.html',
                  {
                      'title': 'About',
                      'message': 'Your application description page.',
                      'year': datetime.now().year,
                  })


import app.test_query_orm as orm


def test_query(request):
    object_list = orm.orm_query1(request)
    return render(request,
                  'app/test_orm.html',
                  {
                      'object_list': object_list,
                      'object_x': None
                  })


def list_inventory_template_for_pm(request):
    """Renders the about page."""
    templateList=PM_Inventory_Template.objects.all()
    template_formset=PMInventoryTemplateFormSet(queryset=templateList)
    context = {'template_formset': template_formset }
    return render(request, 'app/pm_template_json_data.html', context)


@login_required(login_url='login')
@staff_admin_only #@manger_only
def report_incident_bi_prediction(request):
    myuser = request.user
    isNotEmplyQuery = checkEmpyQueryString(request)

    if isNotEmplyQuery:
        my_all_preds =  Prediction_ML2_everity_Incident.objects.filter(incident__inventory__project__company__manager__user=myuser).order_by(
            '-prediction_at')
    else:
        my_all_preds =Prediction_ML2_everity_Incident.objects.filter(id=0 )

    xFilter = PredictionBiSeverityIncidentFilter(request.GET, request=request, queryset=my_all_preds)
    if xFilter.qs.count() > 0:
        predictionList = xFilter.qs

    else:
        predictionList = None


    context = {'predictionList': predictionList, 'predictionFilter': xFilter}
    return render(request, 'app/report_incident_bi_prediction.html', context)

def view_feature_incident_bi_prediction(request,id):

    x_obj = get_object_or_404(Prediction_ML2_everity_Incident, pk=id)
    xyz_open_to_close=(x_obj.incident.incident_close_datetime-x_obj.incident.incident_datetime)
    xyz_open_to_close_hour=round(xyz_open_to_close.total_seconds() / (60*60),1)

    xyz_actual_label= "Critical" if x_obj.incident.incident_severity.id in (1,2) else "Normal"

    context={"item":x_obj,"xyz_open_to_close":xyz_open_to_close,"xyz_actual_label": xyz_actual_label }
    return render(request, 'app/report_incident_feature_detail_bi_prediction.html', context)



