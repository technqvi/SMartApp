import app.advanced_report.report_site_grade as report_director
import app.advanced_report.excel_table_report as ais_table
import app.advanced_report.excel_pivot_report as ais_pivot

import json
from django.shortcuts import render
from django.http import HttpResponse
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime,timedelta
from .forms_report import AdvancedReportSearchForm,ReportDateSearchForm
import app.utility as util
import pandas as pd
from django.contrib import messages

from io import BytesIO
from app.decorators import allowed_users,manger_and_viewer_only,manger_only,staff_admin_only,manger_and_viewer_engineer_only
from app.models import *
from django.contrib.auth.decorators import login_required

from django.shortcuts import render, redirect, get_object_or_404


@login_required(login_url='login')
@manger_only
# @manger_and_viewer_engineer_only
def generate_summarization(request,incident_id):
    import requests
    try:
        x = Incident_Summary.objects.get(incident_id=incident_id)
    except Incident_Summary.DoesNotExist:
        print(f"Incident Summary with ID {incident_id} does not exist")
        print(f"We need to invoke Gen-AI to generate incident summarization")
        response = requests.get(f'http://127.0.0.1:5000/get_incident_summarization_by_id/{incident_id}')
        if response.status_code == 200:
            data = response.json()
            if data['success'] == True:
                try:
                    incident = get_object_or_404(Incident, id=incident_id)
                except Incident.DoesNotExist as exception :
                    messages.error(request, exception)

                incident_summary = data['incident_summarization']
                incident_content = data['incident_content']
                model = data['model']
                incident_updated_at = incident.updated_at

                x = Incident_Summary(incident=incident, incident_updated_at=incident_updated_at, model=model,
                                     input_content=incident_content, output_summary=incident_summary)
                x.save()
            else:
                error_message = f"{response.status_code} : {data['error']}"
                messages.error(request, error_message)


        else:
                messages.error(request, response.status_code)

    context = {"x": x}

    return render(request, 'app/search_summarization.html', context)

@login_required(login_url='login')
@manger_only
# @manger_and_viewer_engineer_only
def search_entry(request):
    context = {}
    return render(request, 'app/search_entry.html', context)



@login_required(login_url='login')
@manger_only
# @manger_and_viewer_engineer_only
def search_result(request,incident_id):
    incident= get_object_or_404(Incident, pk=incident_id)
    detailList=incident.incident_detail_set.all()
    inventory = incident.inventory;

    context = {"incident":incident,"detailList":detailList,"inventory":inventory}
    return render(request, 'app/search_result.html', context)



@login_required(login_url='login')
@staff_admin_only
def report_site_grade(request):

    start_support_param=''
    end_support_param = ''
    year_x=-1
    period_x =-1

    search_form = ReportDateSearchForm(request.POST or None)

    if request.method=='POST':
        year_x=int(request.POST.get('year_list'))
        period_x=int(request.POST.get('quarter_list'))

        start_support_param,end_support_param=util.get_period_selection(int(year_x),int(period_x))
        df_company, df_weight, list_cols = report_director.build_report( start_support_param,end_support_param,)
        title_period=''
        if period_x==0:
            title_period=f"YEAR-{year_x}"
        else:
            title_period = f"Q{period_x}-{year_x}"

        start_support=datetime.datetime.strptime(start_support_param, "%Y-%m-%d").strftime("%d %b %Y")
        end_support= datetime.datetime.strptime(end_support_param, "%Y-%m-%d").strftime("%d %b %Y")

        if 'view_report' in request.POST:
            data = json.loads(df_company.to_json(orient='records'))
            context = {
                'search_form': search_form,
                'period_report': f"{title_period} : {start_support}  - {end_support}",
                'report_cols': list_cols,
                'report_items': data
            }
            return render(request, 'app/report_site_grade.html', context)
        elif  'export_report' in request.POST:
            if len(df_company) > 0 and len(df_weight) > 0:
                try:
                    wb = openpyxl.Workbook(write_only=True)
                    ws_company = wb.create_sheet("Site Grade Summary")
                    for r in dataframe_to_rows(df_company, index=False, header=True, ):
                        ws_company.append(r)

                    ws_wieght = wb.create_sheet("Weight Score Reference")
                    for r in dataframe_to_rows(df_weight, index=False, header=True):
                        ws_wieght.append(r)
                except openpyxl.utils.exceptions.IllegalCharacterError as e:
                    raise e
            buildtime = datetime.datetime.now().strftime('%d%m%y_%H%M')
            file_name = f"SiteGrade_{title_period}_{buildtime}.xlsx"

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={file_name}'

            wb.save(response)
            return response

    else:
        context = {
            'search_form': search_form
        }
    return render(request, 'app/report_site_grade.html', context)


@login_required(login_url='login')
@manger_only
def build_ais_excel_report(request):
    comp_x=''
    start_x=''
    end_x=''

    list_cust_comp = util.list_customer_company(request.user)
    if request.method == 'GET':
        search_form = AdvancedReportSearchForm(request.user)
    elif request.method == 'POST':

        search_form = AdvancedReportSearchForm(request.user,request.POST)
        comp_x= request.POST.get('cust_comp_list')
        start_x=request.POST.get('start_date')
        end_x = request.POST.get('end_date')

        end_x_p1_dt=datetime.datetime.strptime(end_x,'%Y-%m-%d')+timedelta(days=1)
        end_x=end_x_p1_dt.strftime('%Y-%m-%d')

        df_all, dfTableReportDict =ais_table.build_table_report(comp_x,start_x,end_x)
        dfPivotReportDict=ais_pivot.build_pivot_report(df_all)

        if dfTableReportDict is not None and dfPivotReportDict is not None :

            with BytesIO() as b:
                # Use the StringIO object as the filehandle.

              writer = pd.ExcelWriter(b, engine='xlsxwriter')


              comp_obj = Company.objects.get(pk=int(comp_x))
              comp_name = comp_obj.company_name
              buildtime =datetime.datetime.now().strftime('%d%m%y_%H%M')
              start_str =datetime.datetime.strptime(start_x, "%Y-%m-%d").strftime("%d%m%y")
              end_str =datetime.datetime.strptime(end_x, "%Y-%m-%d").strftime("%d%m%y")
              file_name = f"{comp_name}_{start_str}_{end_str}_{buildtime}_report.xlsx"


              for sheet_name, data in dfTableReportDict.items():
                if data.empty==False:
                 data.to_excel(writer, sheet_name=sheet_name,index=False)

              for sheet_name, data in dfPivotReportDict.items():
                 if data.empty==False:
                  data.to_excel(writer, sheet_name=sheet_name)

              writer.save()
                # Set up the Http response.

              response = HttpResponse(
                    b.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
              )
              response['Content-Disposition'] = 'attachment; filename=%s' % file_name
              return response

        else:
            context = {
                'message': "No data to build report"
            }

    return render(request, 'app/report_advance.html',
                  {'search_form': search_form})




